"""
Module 1: SQL-to-File Comparison Routes
Flask Blueprint — handles SQL query, file upload, comparison, results, export.
"""

from flask import Blueprint, request, send_file
import pyodbc
import pandas as pd
import numpy as np
import uuid
import datetime as dt_module
from datetime import timezone
import re

from common.json_utils import (safe_jsonify, sanitize_df_for_json,
                                _is_datetimetz, normalize_timestamp)
from common.db_utils import get_connection_string
from common.storage_manager import save_df, load_df
from common.routes import _touch_activity
from module_1.comparison_engine import run_hybrid_comparison

m1_bp = Blueprint('module_1', __name__)


# ═══════════════════════════════════════════════════════════════════════════
# Helper: Normalize all SQL datetime column types consistently
# ═══════════════════════════════════════════════════════════════════════════

def handle_sql_datetime_types(df):
    """Handle various SQL datetime types consistently.

    Properly formats: DATETIME, DATETIME2, SMALLDATETIME, DATETIMEOFFSET,
    TIME (with/without ms), TIMESTAMP WITH TIME ZONE / LOCAL TIME ZONE.

    Makes a copy of the dataframe to avoid modifying the original.
    """
    df = df.copy()

    for col in df.columns:
        # Detect datetime columns
        if pd.api.types.is_datetime64_any_dtype(df[col]):
            # Check for timezone-aware columns
            if _is_datetimetz(df[col]):
                # Convert to UTC first, then format
                df[col] = df[col].dt.tz_convert('UTC').apply(
                    lambda x: normalize_timestamp(x) if pd.notna(x) else ''
                )
            else:
                # Format with milliseconds if they exist
                df[col] = df[col].apply(
                    lambda x: normalize_timestamp(x) if pd.notna(x) else ''
                )

        # Detect date-only columns
        elif pd.api.types.is_datetime64_any_dtype(df[col]) and not _is_datetimetz(df[col]):
            # For datetime64 columns that aren't timezone-aware, assume they're date-only
            df[col] = df[col].dt.strftime('%Y-%m-%d')

        # Detect time-only columns
        elif pd.api.types.is_timedelta64_dtype(df[col]):
            # For timedelta columns, format as HH:MM:SS
            df[col] = df[col].apply(
                lambda x: normalize_timestamp(x) if pd.notna(x) else ''
            )

        # Detect string columns that might contain datetime info
        elif df[col].dtype == 'object':
            # Try to parse as datetime where possible
            try:
                # First try to parse with timezone handling
                parsed = pd.to_datetime(df[col], errors='ignore', utc=True)
                if pd.api.types.is_datetime64_any_dtype(parsed):
                    if _is_datetimetz(parsed):
                        df[col] = parsed.dt.tz_convert('UTC').apply(
                            lambda x: normalize_timestamp(x) if pd.notna(x) else ''
                        )
                    else:
                        df[col] = parsed.apply(
                            lambda x: normalize_timestamp(x) if pd.notna(x) else ''
                        )
            except Exception:
                # If parsing fails, leave the column as is
                pass

    return df


@m1_bp.route('/api/preview_sql', methods=['POST'])
def preview_sql():
    """
    Executes a SQL query and returns columns + top 5 rows.
    """
    _touch_activity()
    data = request.json
    server = data.get('server')
    database = data.get('database')
    query = data.get('query')
    port = data.get('port')

    if not server or not database or not query:
        return safe_jsonify({"error": "Missing parameters"}, 400)

    # Basic Safety Check (Prevent modifications)
    forbidden_keywords = ['DROP', 'DELETE', 'UPDATE', 'INSERT', 'TRUNCATE', 'ALTER', 'GRANT', 'REVOKE', 'EXEC', 'CREATE', 'MERGE']
    if any(keyword in query.upper() for keyword in forbidden_keywords):
        return safe_jsonify({"error": "Security Alert: Only SELECT queries are permitted in this environment."}, 403)

    conn_str = get_connection_string(server, database, port)

    try:
        conn = pyodbc.connect(conn_str, timeout=10)
        df = pd.read_sql(query, conn)

        # Handle all datetime types consistently
        df = handle_sql_datetime_types(df)

        preview_df = df.head(5)
        columns = list(preview_df.columns)
        rows = sanitize_df_for_json(preview_df).to_dict(orient='records')

        # Include last row for truncated preview display in console
        last_row = None
        if len(df) > 5:
            last_row = sanitize_df_for_json(df.tail(1)).to_dict(orient='records')[0]

        conn.close()
        return safe_jsonify({
            "status": "success",
            "columns": columns,
            "preview_data": rows,
            "row_count_estimate": len(df),
            "last_row": last_row
        })

    except Exception as e:
        return safe_jsonify({"status": "error", "message": str(e)}, 500)


@m1_bp.route('/api/upload_file', methods=['POST'])
def upload_file():
    """
    Uploads a file, saves it temp, and returns columns + preview.
    """
    _touch_activity()
    if 'file' not in request.files:
        return safe_jsonify({"error": "No file part"}, 400)

    file = request.files['file']
    if file.filename == '':
        return safe_jsonify({"error": "No selected file"}, 400)

    unique_id = str(uuid.uuid4())
    filename = file.filename

    try:
        if filename.endswith('.csv'):
            df = pd.read_csv(file)
        elif filename.endswith(('.xls', '.xlsx')):
            df = pd.read_excel(file)
        else:
            return safe_jsonify({"error": "Invalid file format. Only CSV or Excel allowed."}, 400)

        # Handle all datetime types consistently
        df = handle_sql_datetime_types(df)

        # Save to Disk Cache FIRST (Parquet handles NaN natively)
        save_df(df, 'uploads', unique_id)

        # Build preview (sanitize handles NaN/NaT → '' for JSON safety)
        preview_df = df.head(5)
        columns = list(preview_df.columns)
        rows = sanitize_df_for_json(preview_df).to_dict(orient='records')

        return safe_jsonify({
            "status": "success",
            "file_id": unique_id,
            "columns": columns,
            "preview_data": rows,
            "total_rows": len(df)
        })

    except Exception as e:
        return safe_jsonify({"status": "error", "message": str(e)}, 500)


@m1_bp.route('/api/run_comparison', methods=['POST'])
def run_comparison():
    """
    Orchestrates the Comparison:
    1. Fetches FULL SQL data.
    2. Retrieves Uploaded File data from Cache.
    3. Runs Hybrid Comparison Engine.
    4. Caches Result.
    """
    _touch_activity()
    data = request.json
    file_id = data.get('file_id')
    server = data.get('server')
    database = data.get('database')
    query = data.get('query')
    port = data.get('port')
    keys = data.get('keys', [])
    column_mapping = data.get('column_mapping', [])
    file_name = data.get('file_name', 'File')

    if not file_id or not server or not database or not query:
        return safe_jsonify({"error": "Missing required parameters"}, 400)

    # 1. Retrieve File Data
    df_file = load_df('uploads', file_id)
    if df_file is None:
        return safe_jsonify({"error": "File session expired or invalid. Please re-upload."}, 404)

    try:
        # 2. Fetch Full SQL Data
        conn_str = get_connection_string(server, database, port)
        conn = pyodbc.connect(conn_str)
        df_sql = pd.read_sql(query, conn)

        # Handle all datetime types consistently
        df_sql = handle_sql_datetime_types(df_sql)

        conn.close()

        # 3. Apply Column Mapping
        if column_mapping and len(column_mapping) > 0:
            rename_map = {m['file']: m['sql'] for m in column_mapping}
            mapped_sql_cols = [m['sql'] for m in column_mapping]
            mapped_file_cols = [m['file'] for m in column_mapping]

            df_sql = df_sql[[c for c in mapped_sql_cols if c in df_sql.columns]]
            df_file = df_file[[c for c in mapped_file_cols if c in df_file.columns]]
            df_file = df_file.rename(columns=rename_map)

        # 4. Run Logic
        result_df, summary = run_hybrid_comparison(df_sql, df_file, keys, file_name=file_name)

        # Convert all datetime columns to normalized strings
        for col in result_df.columns:
            if (pd.api.types.is_datetime64_any_dtype(result_df[col]) or
                    (hasattr(result_df[col], 'dt') and result_df[col].dt.tz is not None)):
                result_df[col] = result_df[col].apply(
                    lambda x: normalize_timestamp(x) if pd.notna(x) else ''
                )

        # 5. Cache Result
        result_id = str(uuid.uuid4())
        save_df(result_df, 'results', result_id)

        # 6. Return Summary + First Page
        preview_page = sanitize_df_for_json(result_df.head(50)).to_dict(orient='records')

        return safe_jsonify({
            "status": "success",
            "result_id": result_id,
            "summary": summary,
            "preview_rows": preview_page,
            "columns": list(result_df.columns)
        })

    except Exception as e:
        # Log the full error for debugging
        import traceback
        traceback.print_exc()
        return safe_jsonify({"status": "error", "message": str(e)}, 500)


@m1_bp.route('/api/results_page', methods=['GET'])
def get_results_page():
    """
    Pagination for the Data Grid.
    Query Params: result_id, page (1-based), size (default 100)
    """
    result_id = request.args.get('result_id')
    page = int(request.args.get('page', 1))
    size = int(request.args.get('size', 100))

    df = load_df('results', result_id)
    if df is None:
        return safe_jsonify({"error": "Result cache expired. Run comparison again."}, 404)

    # Handle all datetime types consistently
    df = handle_sql_datetime_types(df)

    # Calculate slice
    start = (page - 1) * size
    end = start + size

    if start >= len(df):
        return safe_jsonify({"data": [], "page": page, "has_more": False})

    sliced_df = df.iloc[start:end]
    data = sanitize_df_for_json(sliced_df).to_dict(orient='records')

    return safe_jsonify({
        "data": data,
        "page": page,
        "total_pages": (len(df) // size) + 1,
        "has_more": end < len(df)
    })


@m1_bp.route('/api/export_excel', methods=['GET'])
def export_excel():
    """Generate styled Excel file with color-coded reconciliation results."""
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from io import BytesIO

    result_id = request.args.get('result_id')
    if not result_id:
        return safe_jsonify({"error": "Missing result_id"}, 400)

    df = load_df('results', result_id)
    if df is None:
        return safe_jsonify({"error": "Result cache expired. Run comparison again."}, 404)

    # Handle all datetime types consistently
    df = handle_sql_datetime_types(df)

    # Convert all values to strings for Excel export
    df = df.applymap(lambda x: str(x) if not pd.isna(x) else '')

    # Sanitize all cells for display (handles NaT, time objects, etc.)
    df = sanitize_df_for_json(df)

    # ── Color definitions (match web UI) ──
    header_fill = PatternFill(start_color='334155', end_color='334155', fill_type='solid')
    header_font = Font(color='FFFFFF', bold=True, size=10)
    pre_fill = PatternFill(start_color='FEF9C3', end_color='FEF9C3', fill_type='solid')
    post_fill = PatternFill(start_color='DCFCE7', end_color='DCFCE7', fill_type='solid')
    mismatch_cell_fill = PatternFill(start_color='FECACA', end_color='FECACA', fill_type='solid')
    mismatch_font = Font(bold=True, color='7F1D1D')
    sql_only_fill = PatternFill(start_color='FDE68A', end_color='FDE68A', fill_type='solid')
    file_only_fill = PatternFill(start_color='FECDD3', end_color='FECDD3', fill_type='solid')
    section_title_fill = PatternFill(start_color='F3F4F6', end_color='F3F4F6', fill_type='solid')
    section_font = Font(bold=True, size=11)
    thin_border = Border(
        left=Side(style='thin', color='D1D5DB'),
        right=Side(style='thin', color='D1D5DB'),
        top=Side(style='thin', color='D1D5DB'),
        bottom=Side(style='thin', color='D1D5DB')
    )

    display_cols = [c for c in df.columns if c != '_mismatch_cols']

    # ── Split data by category ──
    mismatched = df[df['status'] == 'Mismatch']
    missing_df = df[df['status'] == 'Only in SQL']
    extra_df = df[df['status'] == 'Only in File']

    wb = Workbook()
    ws = wb.active
    ws.title = "Reconciliation Results"

    def write_section(ws, section_df, title, start_row):
        """Write a section with title, headers, and colored data rows."""
        # Section title row
        ws.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=len(display_cols))
        cell = ws.cell(row=start_row, column=1, value=title)
        cell.font = section_font
        cell.fill = section_title_fill
        start_row += 1

        # Column headers
        for ci, col_name in enumerate(display_cols, 1):
            cell = ws.cell(row=start_row, column=ci, value=col_name)
            cell.fill = header_fill
            cell.font = header_font
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center')
        start_row += 1

        # Data rows
        for _, row in section_df.iterrows():
            mismatch_cols = [c.strip() for c in str(row.get('_mismatch_cols', '')).split(',') if c.strip()]
            status = str(row.get('status', ''))
            source_val = str(row.get('source', ''))

            if status == 'Mismatch':
                base_fill = pre_fill if source_val == 'SQL' else post_fill
            elif status == 'Only in SQL':
                base_fill = sql_only_fill
            elif status == 'Only in File':
                base_fill = file_only_fill
            else:
                base_fill = None

            for ci, col_name in enumerate(display_cols, 1):
                val = row.get(col_name, '')
                cell = ws.cell(row=start_row, column=ci, value=str(val))
                cell.border = thin_border

                if base_fill:
                    cell.fill = base_fill

                if col_name in mismatch_cols and status == 'Mismatch':
                    cell.fill = mismatch_cell_fill
                    cell.font = mismatch_font

            start_row += 1

        return start_row + 1

    current_row = 1

    if len(mismatched) > 0:
        current_row = write_section(ws, mismatched, f"Mismatched Rows ({len(mismatched) // 2})", current_row)
    if len(missing_df) > 0:
        current_row = write_section(ws, missing_df, f"Missing from File / SQL Only ({len(missing_df)})", current_row)
    if len(extra_df) > 0:
        current_row = write_section(ws, extra_df, f"Extra in File / Not in SQL ({len(extra_df)})", current_row)

    if len(df) == 0:
        ws.cell(row=1, column=1, value="No discrepancies found - data matches perfectly!")
        ws.cell(row=1, column=1).font = Font(bold=True, size=12, color='16A34A')

    # Auto-width columns
    for ci, col_name in enumerate(display_cols, 1):
        col_letter = get_column_letter(ci)
        max_len = len(str(col_name))
        for i, (_, row) in enumerate(df.iterrows()):
            if i >= 100:
                break
            val_len = len(str(row.get(col_name, '')))
            if val_len > max_len:
                max_len = val_len
        ws.column_dimensions[col_letter].width = min(max_len + 3, 40)

    # Save to memory and return
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=f'reconciliation_{result_id[:8]}.xlsx'
    )
